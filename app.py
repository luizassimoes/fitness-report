
import re
import textwrap
import numpy as np
import pandas as pd
import seaborn as sns
import streamlit as st
import matplotlib.pyplot as plt
import xml.etree.ElementTree as ET
from intervaltree import IntervalTree
from datetime import date

from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment



def spaced_str(var):
  """Inserts spaces before capitalized letters.

  Args:
    word: The word to be modified. Example: JumpRope

  Returns:
    str: The string with spaces. Example: Jump Rope.
  """

  if isinstance(var, str):
    word = var
    idx_caps = [m.start() for m in re.finditer(r'[A-Z]', word)]

    # Inserts spaces before captalized letters except for the first one
    for i, idx in enumerate(idx_caps[1:]):
      word = word[:idx+i] + ' ' + word[idx+i:]

    return word
  return np.nan


# Building a tree of intervals from df_workout
def build_interval_tree(df_workout):
    # df_workout['startDate'] = pd.to_datetime(df_workout['startDate'])
    # df_workout['endDate']   = pd.to_datetime(df_workout['endDate'])
    tree = IntervalTree()
    for _, row in df_workout.iterrows():
        start = row['startDate'] - pd.Timedelta(seconds=1)
        end = row['endDate'] + pd.Timedelta(seconds=1)
        tree.addi(start, end, row['id'])
    return tree


# Associating IDs using our tree
def assign_workout_id(df, interval_tree):
    df['startDate'] = pd.to_datetime(df['startDate'])
    ids = []
    for time in df['startDate']:
        # Searching for intervals that contain the current time:
        matches = interval_tree[time]
        # Adds the correspondent id:
        ids.append(next(iter(matches)).data if matches else None)
    df['id'] = ids
    return df


def parse_large_xml(file, tag, attribute=None, values=[]):
    """
    Lê um XML grande de forma eficiente, processando elementos específicos.
    
    Args:
    - file: Caminho para o arquivo XML.
    - tag: Tag que você deseja extrair (ex.: 'Record').
    - attribute: Atributo que você deseja filtrar (opcional).
    - values: Valor do atributo que você deseja filtrar (opcional).
    
    Returns:
    - DataFrame contendo os dados extraídos.
    """
    rows = []
    file.seek(0)
    for event, elem in ET.iterparse(file, events=('end',)):
        if elem.tag == tag:

            data = elem.attrib

            if tag == 'Workout':
                calories, distance_km = (0,) * 2
                heart_rate_min, heart_rate_max, heart_rate_avg = (0,) * 3

                for child in elem:
                    if child.tag == 'WorkoutStatistics':
                        if 'ActiveEnergyBurned' in child.attrib.get('type', ''):
                            calories = child.attrib.get('sum')
                        elif 'HeartRate' in child.attrib.get('type', ''):
                            heart_rate_min = child.attrib.get('minimum')
                            heart_rate_max = child.attrib.get('maximum')
                            heart_rate_avg = child.attrib.get('average')
                        elif 'Distance' in child.attrib.get('type', ''):
                            km_sports = ['Running', 'Swimming', 'Cycling', 'Walking', 'Hiking']
                            if data['workoutActivityType'].endswith(tuple(km_sports)):
                                distance_km = child.attrib.get('sum')  # Pega a distância (sum)

                data['calories'] = float(calories)
                data['heart_rate_min'] = heart_rate_min
                data['heart_rate_max'] = heart_rate_max
                data['heart_rate_avg'] = heart_rate_avg
                data['distance_km'] = distance_km

            if attribute and attribute in data:
                if data[attribute] in values:
                    rows.append(data)  # Extrai os atributos como dicionário
            elif not attribute:
                rows.append(data)  # Extrai os atributos como dicionário
            elem.clear()
    return pd.DataFrame(rows)


def tsd(num):
    """ Formats a number including thousands separator. """
    return f"{num:,.0f}".replace(",", ".")

    
def insert_table(sheet, table, dark, light):
    """
    Inserts a table in a Workbook Sheet.

    Args:
    - sheet: The active Workbook where the table should be inserted.
    - table: The table to be inserted.
    - dark: the darked color to use in the table's header and indexes backgrounds.
    - light: the lighter color to use in the table's values backgrounds.
    
    Returns:
    - Sheet containing the solicited table. 
    """

    center_alignment = Alignment(horizontal='center', vertical='center')
    curr_row = sheet.max_row + 1

    # Insere o cabeçalho
    for i_col, col in enumerate(table.columns):
        sheet.cell(row=curr_row, column=i_col+3).value = col
        sheet.cell(row=curr_row, column=i_col+3).fill = PatternFill(start_color=dark, fill_type='solid')
        sheet.cell(row=curr_row, column=i_col+3).font = Font(bold=True)
        if i_col > 0: 
            sheet.cell(row=curr_row, column=i_col+3).alignment = center_alignment
    sheet.cell(row=curr_row, column=2).fill = PatternFill(start_color=dark, fill_type='solid')

    # Insere os dados
    next_row = curr_row + 1
    for i_row in range(table.shape[0]):
        sheet.cell(row=i_row+next_row, column=2).value = i_row+1 # Insere as posições
        sheet.cell(row=i_row+next_row, column=2).fill = PatternFill(start_color=dark, fill_type='solid')
        sheet.cell(row=i_row+next_row, column=2).alignment = center_alignment
        for i_col in range(table.shape[1]):
            value = table.iloc[i_row, i_col]
            sheet.cell(row=i_row+next_row, column=i_col+3).fill = PatternFill(start_color=light, fill_type='solid')
            if i_col > 0:
                sheet.cell(row=i_row+next_row, column=i_col+3).value = value
                sheet.cell(row=i_row+next_row, column=i_col+3).number_format = '#,##0'
                sheet.cell(row=i_row+next_row, column=i_col+3).alignment = center_alignment
            else:
                sheet.cell(row=i_row+next_row, column=i_col+3).value = value
                sheet.cell(row=i_row+next_row, column=i_col+3).alignment = Alignment(vertical='center', wrap_text=True)

    return sheet


st.set_page_config(page_title="Fitness Wrapped", page_icon="💪")
st.title("🏋️ Fitness Wrapped")

st.write("Here is how you can use this app:   (if you already know it, just [skip to the importer](#importer))")
st.write("1. On your iPhone, open the Health app.")
st.write("2. Tap your profile picture (or initials) in the top right corner.")
st.write("3. Scroll down the page and click 'Export All Health Data'")
st.write("4. Confirm and wait a little bit for the process to complete.")
st.write("5. Once finished, locate the zip file and extract its contents.")
st.write("6. Open the extracted folder and navigate to: apple_health_export > export.xml")
st.write("7. That's the file we need: export.xml.")
st.write("8. Upload the file below and enjoy exploring how your year went!")


st.markdown('<div id="importer"></div>', unsafe_allow_html=True)
today = date.today()
this_year  = today.year
this_month = today.month

year_range    = list(range(this_year, 2000, -1))
default_year  = this_year if this_month == 12 else this_year - 1
default_index = year_range.index(default_year)
selected_year = st.selectbox("Choose a year:", year_range, index=default_index)

my_file = st.file_uploader("Select a file", type=["xml"], label_visibility="hidden")

if my_file is not None:
    st.write("Importing Fitness data...")
    try:
        df_workout = parse_large_xml(my_file, tag='Workout')
        df_workout = df_workout.drop(columns=['durationUnit', 'sourceName', 'sourceVersion'], axis=1)
        df_workout = df_workout[df_workout['startDate'].str.startswith(f"{selected_year}")].reset_index(drop=True)
    except:
        st.error("Are you sure you uploaded the correct file? Something went wrong... please try again.")
        st.stop()

    if df_workout.empty:
        st.error("Sorry, there are no Workout records for that year.")
        st.stop()

    wk_int_columns = ['duration', 'heart_rate_min', 'heart_rate_max', 'heart_rate_avg', 'calories']
    df_workout[wk_int_columns] = df_workout[wk_int_columns].apply(lambda row: pd.to_numeric(row, errors='coerce').fillna(0).astype(int))

    df_workout['workoutActivityType'] = df_workout['workoutActivityType'].str.split('ActivityType').where(df_workout['workoutActivityType'].notnull(), np.nan).str[1]
    df_workout['workoutActivityType'] = df_workout['workoutActivityType'].apply(spaced_str)

    date_columns = ['startDate', 'endDate', 'creationDate'] 
    for col in date_columns:
        df_workout[col] = pd.to_datetime(df_workout[col])


    if all(df_workout['heart_rate_avg'] == 0):
        st.write("Importing Health data...")

        att_list = ['HKQuantityTypeIdentifierHeartRate'] 
        df_heart_rate = parse_large_xml(my_file, tag='Record', attribute='type', values=att_list)
        df_heart_rate = df_heart_rate.drop(['sourceName', 'sourceVersion', 'device', 'unit', 'creationDate', 'endDate'], axis=1)
        df_heart_rate = df_heart_rate.rename(columns={'value': 'heart_rate'})
        df_heart_rate = df_heart_rate[df_heart_rate['startDate'].str.startswith(f"{selected_year}")].reset_index(drop=True)
        
        df_heart_rate[['heart_rate']] = df_heart_rate[['heart_rate']].apply(lambda row: pd.to_numeric(row, errors='coerce').fillna(0).astype(int))

        df_workout = df_workout.drop(columns=['heart_rate_min', 'heart_rate_max', 'heart_rate_avg'])

        # Processo de união dos dataframes:
        ids = pd.Series(range(1, len(df_workout) + 1))
        df_workout = pd.concat([ids.rename('id'), df_workout], axis=1)

        workout_tree = build_interval_tree(df_workout)
        df_heart_rate = assign_workout_id(df_heart_rate, workout_tree)  

        df_heart_rate = df_heart_rate[~df_heart_rate['id'].isna()].reset_index(drop=True)

        df_heart_rate = df_heart_rate.groupby('id').agg(
                    heart_rate_avg=('heart_rate', 'mean'),
                    heart_rate_max=('heart_rate', 'max'),
                    heart_rate_min=('heart_rate', 'min')
                    ).reset_index()

        hr_int_columns = [col for col in df_heart_rate.columns if col.startswith('heart_rate')]
        df_heart_rate[hr_int_columns] = df_heart_rate[hr_int_columns].apply(lambda row: pd.to_numeric(row, errors='coerce').fillna(0).astype(int))

        df_workout = pd.merge(df_workout, df_heart_rate, on='id', how='left')


    st.write("Importing Activity data...")
    df_activity = parse_large_xml(my_file, tag='ActivitySummary')
    df_activity = df_activity[['dateComponents', 'activeEnergyBurned', 'activeEnergyBurnedGoal', 'appleExerciseTime']]
    df_activity = df_activity[df_activity['dateComponents'].str.startswith(f"{selected_year}")].reset_index(drop=True)

    ac_int_columns = ['activeEnergyBurned', 'appleExerciseTime', 'activeEnergyBurnedGoal']
    df_activity[ac_int_columns] = df_activity[ac_int_columns].apply(lambda row: pd.to_numeric(row, errors='coerce').fillna(0).astype(int))


    # st.write("All data imported.")
    st.write("Calculating your results...")

    # Adjusting numeric values
    st.write()
    
    
    # Adding date columns:
    months_dict = {1: 'JAN', 2: 'FEB', 3: 'MAR', 4: 'APR', 5: 'MAY', 6: 'JUN', 
                7: 'JUL', 8: 'AUG', 9: 'SEP', 10: 'OCT', 11: 'NOV', 12: 'DEC'}

    df_workout['date']  = df_workout['startDate'].dt.date
    df_workout['month'] = df_workout['startDate'].dt.month
    df_workout['year']  = df_workout['startDate'].dt.year
    df_workout['month_name'] = df_workout['month'].map(months_dict)

    df_activity['dateComponents'] = pd.to_datetime(df_activity['dateComponents'])
    df_activity['month'] = df_activity['dateComponents'].dt.month
    df_activity['month_name'] = df_activity['month'].map(months_dict)


    # Results:

    # Distance:
    df_workout['distance_km'] = df_workout['distance_km'].astype(float).astype(int)
    kms = df_workout[~df_workout['distance_km'].isna()].groupby('workoutActivityType').agg({'duration': 'sum', 'distance_km': 'sum', 'creationDate': 'count'}).reset_index()
    kms = kms[kms['distance_km'] != 0]
    kms = kms.sort_values('distance_km', ascending=False)

    total_kms = kms['distance_km'].sum()
    total_kms = int(total_kms)


    # Top sports:
    sports_by_count    = df_workout['workoutActivityType'].value_counts()
    total_sports       = len(sports_by_count)
    top_sport_by_count = sports_by_count.idxmax()
    count_top_sport    = sports_by_count.max()

    top_sports_by_time = df_workout.groupby(by=['workoutActivityType']).agg({'duration': 'sum', 'calories': 'sum', 'date': 'count'}).sort_values(by='duration', ascending=False)

    top_sport_by_time       = top_sports_by_time.idxmax().iloc[0]
    time_top_sport          = top_sports_by_time.max().iloc[0]
    count_top_sport_by_time = sports_by_count[top_sport_by_time]

    top_5_sports_by_time    = top_sports_by_time.iloc[:5, :].reset_index()
    top_5_sports_by_time.columns = ['Sport', 'Total Minutes', 'Total Calories', 'Sessions']

    
    # Heart Rate:
    highest_heart_rate = df_workout.loc[df_workout['heart_rate_max'].idxmax(), :]
    heart_rate_per_sport = df_workout.groupby(['workoutActivityType']).agg({'heart_rate_avg': 'mean'})
    sport_highest_avg_heart_rate = heart_rate_per_sport.idxmax().iloc[0]
    highest_avg_heart_rate = int(heart_rate_per_sport.max().iloc[0])

    
    # Calories:
    calories_per_sport = df_workout.groupby(['workoutActivityType']).agg({'calories': 'mean'})
    sport_highest_avg_calories = calories_per_sport.idxmax().iloc[0]
    highest_avg_calories = int(calories_per_sport.max().iloc[0])

    
    # Total time exercising and goals:
    exercise_total_time = top_sports_by_time['duration'].sum()
    exercise_time_per_day = int(exercise_total_time/365)
    exercise_total_days = len(df_workout['date'].drop_duplicates())
    exercise_total_calories = top_sports_by_time['calories'].sum()
    exercise_calories_per_day = int(exercise_total_calories/365)

    
    energy_goal_check = df_activity[(df_activity['activeEnergyBurned'] >= df_activity['activeEnergyBurnedGoal']) & (df_activity['activeEnergyBurnedGoal'] > 0)].reset_index(drop=True)
    days_energy_goal_check = energy_goal_check.shape[0]

    percent_goal_check_so_far = round(days_energy_goal_check/365*100, 2)

    
    # Top day of exercises:
    top_exercise_day_data = df_workout.groupby(['date']).agg({'duration': 'sum', 'calories': 'sum'})

    top_exercise_day = top_exercise_day_data.idxmax()['duration']

    top_exercise_day_calories = top_exercise_day_data.loc[top_exercise_day, 'calories']
    top_exercise_day_time = top_exercise_day_data.loc[top_exercise_day, 'duration']

    top_exercise_day_training = df_workout.loc[df_workout['date'] == (top_exercise_day)]
    top_exercise_day_training = top_exercise_day_training.rename(
        columns={
            'workoutActivityType': 'Sport',
            'duration': 'Total Minutes',
            'calories': 'Total Calories'
        }
    )

    top_exercise_day = pd.Timestamp(top_exercise_day)

    
    # Most active day:
    top_active_day_data = df_activity.loc[df_activity['activeEnergyBurned'].idxmax(), ['dateComponents', 'activeEnergyBurned', 'appleExerciseTime']]
    top_active_day = top_active_day_data['dateComponents']
    top_active_day_calories = top_active_day_data['activeEnergyBurned']
    top_active_day_time = top_active_day_data['appleExerciseTime']

    top_active_day_training = df_workout.loc[df_workout['date'] == (top_active_day)]
    top_active_day_training = top_active_day_training.rename(
        columns={
            'workoutActivityType': 'Sport',
            'duration': 'Total Minutes',
            'calories': 'Total Calories'
        }
    )

    top_active_day_training_none = None

    
    # Showing all of the results:

    
    row_year_1 = f'You exercised for {tsd(exercise_total_time)} minutes this year in {exercise_total_days} different days and burned {tsd(exercise_total_calories)} calories!'
    row_year_2 = f'That is an average of {exercise_time_per_day} minutes and {exercise_calories_per_day} calories per day.'

    row_mileage_1   = f'You covered a lot of kilometers this year... {tsd(total_kms)}!'
    row_mileage_2   = 'Check it out how you did it:'
    table_mileage_1 = kms.copy()
    table_mileage_1.columns = ['Sport', 'Total Minutes', 'Total kms', 'Sessions']

    row_sports_1 = f'And you experimented {total_sports} different sports!'
    row_sports_2 = f'Your top sport by number of registers is: {top_sport_by_count} and you practiced it {count_top_sport} times this year.'

    if top_sport_by_count != top_sport_by_time:
        row_sports_3 = f'But your top sport by time is: {top_sport_by_time} and you practiced it {count_top_sport_by_time} times this year for {tsd(time_top_sport)} minutes.'
    else:
        row_sports_3 = f'And your top sport by time is also: {top_sport_by_time}! You practiced it for {tsd(time_top_sport)} minutes.'

    row_sports_4 = f"You burned an average of {int(top_5_sports_by_time.loc[0, 'Total Calories']/count_top_sport_by_time)} calories per {top_sport_by_time} session."

    row_sports_5 = f'Your top 5 sports are:'
    table_sports_1 = top_5_sports_by_time

    
    row_heart_1 = f'The sport that makes your heart race is: {sport_highest_avg_heart_rate}!'
    row_heart_2 = f'The average heart rate you get when practicing it is {highest_avg_heart_rate} bpm.'

    
    row_calories_1 = f'But the sport that really makes you burn is {sport_highest_avg_calories} with an average of {highest_avg_calories} calories per session!'

    
    row_day_1 = f'The day you exercised the most was {top_exercise_day.month_name()} {top_exercise_day.day}.'
    row_day_2 = f'You worked out for {top_exercise_day_time} minutes and burned {int(top_exercise_day_calories)} calories!'

    row_day_3 = f'Exercises that day:'
    top_exercise_day_training = top_exercise_day_training.reset_index(drop=True)
    table_day_1 = top_exercise_day_training[['Sport', 'Total Minutes', 'Total Calories']]

    
    row_day_4 = f'Now, your most active day was... {top_active_day.month_name()} {top_active_day.day}!'
    row_day_5 = f'Your movement ring reached {tsd(int(top_active_day_calories))} calories and your exercise ring registered {top_active_day_time} minutes!'

    if top_active_day_training.empty:
        row_day_6 = 'However, there are no exercises registered for this day...'
        table_day_2 = ''
    else:
        top_active_day_training = top_active_day_training.reset_index(drop=True)
        row_day_6 = f"Exercises that day: "
        table_day_2 = top_active_day_training[['Sport', 'Total Minutes', 'Total Calories']]

    
    row_goals_1 = f'You reached your movement goal on {days_energy_goal_check} days this year!'
    row_goals_2 = f'That represents {percent_goal_check_so_far}% of the year!'

    row_view_1 = f'Below you can see your results throughout the year.'

    
    df_activity_per_month = df_activity.groupby(['month', 'month_name']).agg({'appleExerciseTime': 'sum'}).reset_index()
    df_activity_per_month['hours'] = df_activity_per_month['appleExerciseTime'].apply(lambda value: round(value/60, 1))

    
    fig, ax1 = plt.subplots(figsize=(12, 5))

    mean_value = df_activity_per_month['hours'].mean()
    ax1.axhline(mean_value, color='orange', linestyle=':', linewidth=2, label='Average')
    ax1.annotate(
        f'{mean_value:.2f} h',  # Texto a ser exibido, ajustado para duas casas decimais
        xy=(0.93, mean_value*1.02),  # Coordenadas (x, y) para posicionar a anotação
        xycoords=('axes fraction', 'data'),  # Define que x usa fração dos eixos e y os dados
        fontsize=10,  # Tamanho da fonte
        color='orange',  # Cor do texto
        bbox=dict(boxstyle="round,pad=0.3", edgecolor='none', facecolor='white', alpha=0.7)  # Fundo branco com leve transparência
    )

    lp_time_per_month = sns.lineplot(data=df_activity_per_month, x='month_name', y='hours')
    ax1.set_xlabel('');
    ax1.yaxis.set_visible(False)

    y_min = df_activity_per_month['hours'].min()
    y_max = df_activity_per_month['hours'].max()
    ax1.set_ylim(y_min * 0.8, y_max * 1.08)
    ax1.set_title('Total Workout Time per Month')

    for i, row in df_activity_per_month.iterrows():
        ax1.annotate(
            text=f"{str(row['hours']).replace('.', ',')} h",  # Texto da anotação
            xy=(row['month_name'], row['hours']),  # Coordenadas do ponto
            xytext=(0, 5),  # Deslocamento em relação ao ponto (x, y)
            textcoords='offset points',  # Usa o deslocamento definido acima
            ha='center', fontsize=10, color='black', bbox=dict(facecolor='white', alpha=0.6, edgecolor='none')
        )


    img_buffer_1 = BytesIO()
    plt.tight_layout()
    plt.savefig(img_buffer_1, format='png', dpi=100)

    
    df_workout_gb = df_workout.groupby(['month', 'month_name', 'workoutActivityType']).agg({'duration': 'sum'}).reset_index()
    df_workout_sport_time = df_workout.groupby(['workoutActivityType']).agg({'duration': 'sum'}).reset_index()
    df_workout_sport_time = df_workout_sport_time.sort_values(by='duration', ascending=False)

    fig, ax1 = plt.subplots(figsize=(18, 5))
    lp_time_per_sport = sns.barplot(data=df_workout_sport_time, x='workoutActivityType', y='duration', hue='workoutActivityType');
    ax1.yaxis.set_visible(False)
    ax1.set_ylim()

    labels = [textwrap.fill(label, width=11) for label in df_workout_sport_time['workoutActivityType']]  # Limitando a 14 caracteres e quebrando a linha
    ax1.set_xticks(range(len(labels)))  # Definindo o número de ticks com base no número de rótulos
    ax1.set_xticklabels(labels, rotation=0, fontdict={'fontsize': 10});  
    ax1.set_xlabel('');
    ax1.set_title('Your Workout Distribution', fontdict={'fontsize': 20})

    y_max = df_workout_sport_time['duration'].max()
    ax1.set_ylim(0, y_max * 1.08)

    for container in lp_time_per_sport.containers:
        labels = [f'{v.get_height():,.0f} min'.replace(',', '.') if v.get_height() != 0 else '' for v in container]
        lp_time_per_sport.bar_label(container, labels=labels, padding=3, fontsize=10, bbox=dict(boxstyle="round,pad=0.2", facecolor='white', alpha=0, edgecolor='none'))

    img_buffer_2 = BytesIO()
    plt.tight_layout()
    plt.savefig(img_buffer_2, format='png', dpi=100)

    
    # Gerando o relatório:
    wb = Workbook()
    sheet = wb.active

    sheet.sheet_view.showGridLines = False
    sheet.title = 'Fitness'

    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 4
    for col in ['C', 'D', 'E', 'F']:
        sheet.column_dimensions[col].width = 16

    sheet.cell(row=1, column=1).value = 'WORKOUT REPORT'
    sheet.cell(row=1, column=1).font = Font(size=28, name='Bahnschrift Light')

    sheet.cell(row=3, column=1).value = 'YOUR INFO:'
    sheet.cell(row=3, column=1).font = Font(size=14, name='Bahnschrift Light')


    dark_orange  = 'F1A983'
    light_orange = 'FDF0E9'
    dark_blue    = 'A6C9EC'
    light_blue   = 'EBF3FB'
    center_alignment = Alignment(horizontal='center', vertical='center')

    sheet.row_dimensions[4].height = 9

    # Year general info:
    sheet.cell(row=5, column=1).value     = selected_year
    sheet.cell(row=5, column=1).fill      = PatternFill(start_color='B5E6A2', fill_type='solid')
    sheet.cell(row=5, column=1).alignment = center_alignment
    sheet.cell(row=5, column=2).value     = row_year_1
    sheet.cell(row=6, column=2).value     = row_year_2

    # Mileage
    curr_row = sheet.max_row + 2

    sheet.cell(row=curr_row, column=1).value     = 'Mileage'
    sheet.cell(row=curr_row, column=1).fill      = PatternFill(start_color='4EB6B1', fill_type='solid')
    sheet.cell(row=curr_row, column=1).alignment = center_alignment
    sheet.cell(row=curr_row, column=2).value     = row_mileage_1
    sheet.cell(row=curr_row+2, column=2).value   = row_mileage_2
    sheet.cell(row=curr_row+2, column=2).font    = Font(bold=True)
    sheet = insert_table(sheet, table_mileage_1, '4EB6B1', 'DEF2F2')

    # Sports
    curr_row = sheet.max_row + 2

    sheet.cell(row=curr_row, column=1).value     = 'Sports'
    sheet.cell(row=curr_row, column=1).fill      = PatternFill(start_color=dark_orange, fill_type='solid')
    sheet.cell(row=curr_row, column=1).alignment = center_alignment
    sheet.cell(row=curr_row, column=2).value     = row_sports_1
    sheet.cell(row=curr_row+1, column=2).value   = row_sports_2
    sheet.cell(row=curr_row+2, column=2).value   = row_sports_3
    sheet.cell(row=curr_row+3, column=2).value   = row_sports_4
    sheet.cell(row=curr_row+5, column=2).value   = row_sports_5
    sheet.cell(row=curr_row+5, column=2).font    = Font(bold=True)
    sheet = insert_table(sheet, table_sports_1, dark_orange, light_orange)

    # Heart
    curr_row = sheet.max_row + 2

    sheet.cell(row=curr_row, column=1).value     = 'Heart'
    sheet.cell(row=curr_row, column=1).fill      = PatternFill(start_color='E49EDD', fill_type='solid')
    sheet.cell(row=curr_row, column=1).alignment = center_alignment
    sheet.cell(row=curr_row, column=2).value     = row_heart_1
    sheet.cell(row=curr_row+1, column=2).value   = row_heart_2

    # Calories
    curr_row = sheet.max_row + 2

    sheet.cell(row=curr_row, column=1).value     = 'Calories'
    sheet.cell(row=curr_row, column=1).fill      = PatternFill(start_color='F5B20B', fill_type='solid')
    sheet.cell(row=curr_row, column=1).alignment = center_alignment
    sheet.cell(row=curr_row, column=2).value     = row_calories_1

    # Days
    curr_row = sheet.max_row + 2

    sheet.cell(row=curr_row, column=1).value     = 'Days'
    sheet.cell(row=curr_row, column=1).fill      = PatternFill(start_color=dark_blue, fill_type='solid')
    sheet.cell(row=curr_row, column=1).alignment = center_alignment
    sheet.cell(row=curr_row, column=2).value     = row_day_1
    sheet.cell(row=curr_row+1, column=2).value   = row_day_2
    sheet.cell(row=curr_row+3, column=2).value   = row_day_3
    sheet.cell(row=curr_row+3, column=2).font    = Font(bold=True)
    sheet = insert_table(sheet, table_day_1, dark_blue, light_blue)

    curr_row = sheet.max_row + 3

    sheet.cell(row=curr_row, column=2).value   = row_day_4
    sheet.cell(row=curr_row+1, column=2).value = row_day_5
    sheet.cell(row=curr_row+3, column=2).value = row_day_6
    if not row_day_6.startswith('However'):
        sheet = insert_table(sheet, table_day_2, dark_blue, light_blue)

    # Goals
    curr_row = sheet.max_row + 2

    sheet.cell(row=curr_row, column=1).value     = 'Goals'
    sheet.cell(row=curr_row, column=1).fill      = PatternFill(start_color='E76B91', fill_type='solid')
    sheet.cell(row=curr_row, column=1).alignment = center_alignment
    sheet.cell(row=curr_row, column=2).value     = row_goals_1
    sheet.cell(row=curr_row+1, column=2).value   = row_goals_2

    # View
    curr_row = sheet.max_row + 2

    sheet.cell(row=curr_row, column=1).value     = 'View'
    sheet.cell(row=curr_row, column=1).fill      = PatternFill(start_color='FCEB6A', fill_type='solid')
    sheet.cell(row=curr_row, column=1).alignment = center_alignment
    sheet.cell(row=curr_row, column=2).value     = row_view_1

    img_1 = Image(img_buffer_1)
    img_1.width = 730
    img_1.height = 305
    img_2 = Image(img_buffer_2)
    img_2.width = 730
    img_2.height = 205

    row_img1 = curr_row+3
    sheet.add_image(img_1, f'B{row_img1}')
    sheet.add_image(img_2, f'B{row_img1+16}')

    excel_title = f"{selected_year}_fitness_report.xlsx"

    
    # wb.save(excel_title)
    output = BytesIO()
    wb.save(output)


    col1, col2, col3 = st.columns(3)
    with col2:
        st.download_button(
            label="Download my Fitness Report",
            data=output,
            file_name=excel_title,
            mime='application/xlsx' #"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Please, submit a file to get started.")
